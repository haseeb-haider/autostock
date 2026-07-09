"""Builds an .xlsx export of the current data: ingredients, products & recipes,
suppliers, the purchases log, and the sales log - one sheet each."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config
from .inventory import snapshot
from .models import Product, Purchase, Sale, Supplier

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_LABELS = {"ok": "In Stock", "reorder": "Low Stock", "out": "No Stock"}


def _sheet(wb, title, headers):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    return ws


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 3
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 42)


def _money_columns(ws, *col_letters):
    fmt = f'"{config.CURRENCY_SYMBOL}"#,##0.00'
    for letter in col_letters:
        for cell in ws[letter][1:]:
            cell.number_format = fmt


def build_workbook(db):
    wb = Workbook()
    wb.remove(wb.active)

    rows = snapshot(db)
    ws = _sheet(wb, "Ingredients List", [
        "Ingredient Name", "Unit of Measure", "Supplier", "Cost per Unit",
        "Reorder Level", "Current Qty in Stock", "Status", "Inventory Value",
    ])
    for r in rows:
        ws.append([r["name"], r["unit"], r["supplier"], r["cost_per_unit"],
                   r["reorder_point"], r["stock"], STATUS_LABELS[r["status"]], r["value"]])
    _money_columns(ws, "D", "H")
    _autosize(ws)

    ws = _sheet(wb, "Products & Recipe List", [
        "Product Name", "Retail Price", "Ingredient", "Qty per Unit Sold",
    ])
    for p in db.query(Product).all():
        if p.recipe_items:
            for item in p.recipe_items:
                ws.append([p.name, p.retail_price,
                           item.ingredient.name if item.ingredient else "", item.qty_per_unit])
        else:
            ws.append([p.name, p.retail_price, "", ""])
    _money_columns(ws, "B")
    _autosize(ws)

    ws = _sheet(wb, "Supplier List", ["Supplier Name", "Email", "Lead Time (days)"])
    for s in db.query(Supplier).all():
        ws.append([s.name, s.email, s.lead_time_days])
    _autosize(ws)

    suppliers_by_id = {s.id: s.name for s in db.query(Supplier).all()}
    ws = _sheet(wb, "Supplier Purchases Log", [
        "Order Date", "Ingredient Name", "Quantity", "Unit Cost", "Total Cost",
        "Supplier", "Delivery Status", "Delivery Date",
    ])
    for p in db.query(Purchase).order_by(Purchase.order_date).all():
        ws.append([
            p.order_date, p.ingredient.name if p.ingredient else "", p.qty, p.unit_cost,
            round((p.qty or 0) * (p.unit_cost or 0), 2),
            suppliers_by_id.get(p.supplier_id, ""), p.status.capitalize(), p.delivered_date,
        ])
    _money_columns(ws, "D", "E")
    _autosize(ws)

    ws = _sheet(wb, "Customer Sales Log", [
        "Order Date", "Product Name", "Quantity", "Price per Unit", "Gross Sales",
    ])
    for s in db.query(Sale).order_by(Sale.sale_date).all():
        ws.append([s.sale_date, s.product.name if s.product else "", s.qty, s.unit_price,
                   round((s.qty or 0) * (s.unit_price or 0), 2)])
    _money_columns(ws, "D", "E")
    _autosize(ws)

    return wb
